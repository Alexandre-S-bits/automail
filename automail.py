import os
import smtplib
from getpass import getpass
from email import encoders
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from time import sleep
from openpyxl import load_workbook
from datetime import datetime


# Servidor do Gmail
host_server = 'smtp.gmail.com'
port_server = '587'


# Conectar e iniciar o servidor com seguranca TLS
server = smtplib.SMTP(host_server, port_server)
server.ehlo()
server.starttls()


def realizar_login():
    """Pede ao usuario para inserir o email e a senha para fazer login"""
    print(f' INFORMAÇÕES PARA LOGIN')
    while True:
        email_login = input(' Email: ')
        password_login = getpass(prompt=' Senha: ')
        try:
            server.login(email_login, password_login)
            print(f'\n Sucesso ao logar!')
            print(f'{"="*80}\n')
            break
        except:
            print(' Email ou senha incorretas!\n')
    return email_login, password_login
        

def adicionar_anexo():
    anexos = []
    while True:
        caminho_arquivo = input(' Informe [0] para voltar ou digite o caminho do arquivo: ')
        if caminho_arquivo == '0':
            print(f' Quantidade de anexos: {len(anexos)}')
            print(f' Lista de anexos adicionados: ')
            [print(anexo) for anexo in anexos]
            while True:
                opcao = input(' Prosseguir com essa lista de anexos?\n'
                              ' N - Não anexar nenhum arquivo\n'
                              ' C - Confirmar lista de anexos e prosseguir\n'
                              ' Z - Limpar lista e anexar novos arquivos\n'
                              ' >>> ').upper()
                if opcao == 'C':
                    return anexos
                elif opcao == 'Z':
                    anexos.clear()
                elif opcao == 'N':
                    anexos.clear()
                    return None
                print(f'A opcao {opcao} é invalida')
        try:
            ler_arquivo = open(caminho_arquivo, mode='rb')
            anexos.append(ler_arquivo)
            ler_arquivo.close()
        except FileNotFoundError:
            print(f' O arquivo {caminho_arquivo} não existe')


def converter_anexos():

    lista_anexos = adicionar_anexo()
    if lista_anexos is None:
        return None
    for anexo in lista_anexos:
        ler_anexo = open(f'{anexo}', mode='rb')

        # lendo arquvios em forma binario e encondando para base 64
        anexo = MIMEBase('application', 'octet-stream')
        anexo.set_payload(ler_anexo.read())
        encoders.encode_base64(anexo)

        # adicionando o cabeçalho do arquivo
        nome_anexo = os.path.basename(str(anexo))
        anexo.add_header('Content-Disposition', f'attachment; filename={nome_anexo}')
        msg_email.attach(anexo)
        ler_anexo.close()
    return anexo


def converter_conteudo(arquivo):
    try:
        with open(arquivo, mode='r', encoding='utf-8') as arquivo_modelo:
            conteudo_base = arquivo_modelo.read()
    except FileNotFoundError:
        print(f' O arquivo {arquivo_modelo} não foi encontrado')
        
    string = conteudo_base.replace('${nome_destinario}', str(guia[f'C{cont}'].value))
    string = string.replace('${cpnj_cpf_destinatario}', str(guia[f'E{cont}'].value))
    string = string.replace('${informacao_1}', str(guia[f'H{cont}'].value))
    string = string.replace('${informacao_2}', str(guia[f'I{cont}'].value))
    string = string.replace('${informacao_3}', str(guia[f'J{cont}'].value))
    string = string.replace('${informacao_4}', str(guia[f'K{cont}'].value))
    string = string.replace('${informacao_5}', str(guia[f'L{cont}'].value))
    return string


def destinatarios():
    string_emails = str(guia[f'D{cont}'].value)
    lista_destinatarios = string_emails.replace(' ', '')
    if ',' in string_emails:
        lista_destinatarios = string_emails.split(',')
        lista_destinatarios = ', '.join(lista_destinatarios)
    return lista_destinatarios


print(f'\n{" A U T O M A I L ":=^80}\n')
# logar no servidor
user, password = realizar_login()
server.login(user, password)
planilha = ''

while True:
    print(f' SELECIONAR ARQUIVOS')
    # arquivo html base
    arquivo_base = input(' Caminho do arquivo base do corpo do email: ')

    # lendo planilha
    caminho_planilha = input(' Caminho da planilha: ')

    # nome da guia da planilha
    nome_guia = input(' Nome da guia na planilha: ')

    try:
        planilha = load_workbook(caminho_planilha, data_only=True)  # ler somente os dados da planilha
        guia = planilha[nome_guia]  # abre na guia especificada dentro dos colchetes
        if '.html' in arquivo_base and nome_guia in planilha.sheetnames and ('.xls' in caminho_planilha or '.xlsm' in caminho_planilha):
            print(f'{"="*80}\n')
            break
    except:
        print(' Informaçoes incorretas! Digite novamente\n')


# contador para alterar de linha
cont = 2
while True:
    nome_destinario = guia[f'C{cont}'].value
    if nome_destinario is None or nome_destinario == '' or nome_destinario == 'NÃO CADASTRADO':
        break
    msg_email = MIMEMultipart()
    msg_email['From'] = user
    email_destinatarios = destinatarios()
    msg_email['To'] = email_destinatarios
    msg_email['Subject'] = str(guia[f'F{cont}'].value)

    corpo_email = converter_conteudo(arquivo_base)
    msg_email.attach(MIMEText(corpo_email, 'html'))
    
    # verifica se terá anexo
    if guia[f'G{cont}'].value != 'NÃO':
        while True:
            add_anexo = input(' Deseja adicionar anexos? [S/N] ').upper()
            if add_anexo == 'N':
                break
            elif add_anexo == 'S':
                arquivos_anexos = converter_anexos()
                if arquivos_anexos is not None:
                    converter_anexos()
                break
            else:
                print(f' A Opcao {add_anexo} é invalida')

    sleep(3)
    server.sendmail(msg_email['From'], msg_email['To'], msg_email.as_string())
    sleep(2)

    horario_atual = datetime.now().strftime('%d/%m/%Y %H:%M:%S') 
    # Escreve o horário atual na coluna L da planilha
    guia[f'A{cont}'] = str(horario_atual)
    planilha.save(caminho_planilha)
    sleep(0.2)
    print(f' Email enviado com sucesso\n Codigo: {str(guia[f"B{cont}"].value)}\n Assunto: {str(guia[f"F{cont}"].value)}\n Destinario(s): {nome_destinario}\n Email(s): {str(guia[f"D{cont}"].value)}\n___________________________________________\n')
    cont += 1

planilha.close()
print(' Fechando servidor de email...')
server.quit()
input(' Servidor finalizado. Aperte qualquer tecla para sair ')
